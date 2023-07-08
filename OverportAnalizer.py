from types import NoneType

import input_data
import shutil
import openpyxl
from colorama import Fore
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

import regulars


class OverportAnalizer:
    PATH = 'input_data/ads_journal.xlsx'

    DATE = 3
    ADDRESS = 4
    TASK = 5
    STATUS = 9
    COMMENT = 6

    def __init__(self):

        self.wb: Workbook = openpyxl.load_workbook(self.PATH, read_only=False)
        self.wb.create_sheet('Without_Duplicates', 0)
        self.wb.save(self.PATH)
        self.sheet: Worksheet = self.wb.active
        self.sheet_origin: Worksheet = self.wb['Лист 1']
        self.remove_duplicates()

    def remove_duplicates(self):
        print('Removing duplicates...')
        moved_row_count = 0
        for task in range(2, 30):
            if self.sheet_origin[task][self.COMMENT].value is None:
                # self.sheet.append(self.sheet_origin.iter_rows(min_row=task, max_row=task, values_only=True))

                moved_row_count += 1
            else:
                for word in regulars.DUPLICATES:
                    if str(self.sheet_origin[task][self.COMMENT].value).lower().find(word) == -1:
                        self.sheet.append(self.sheet_origin.iter_rows(min_row=task, max_row=task, values_only=True))
                        moved_row_count += 1
        self.wb.save(self.PATH)
        print(f'Удалено {self.sheet_origin.max_row - moved_row_count - 1} дублей')

        # ws1.append(row) for row in ws2.iter_rows(min_row=2, max_row=2, values_only=True)

    def show_tasks(self, task_quantity=all, date_range_min='2020-01-01', date_range_max='2100-01-01'):
        # Определяем нужное кол-во тасков
        if task_quantity == all:
            tq = self.sheet.max_row
        else:
            tq = task_quantity

        # Определяем временно диапазон
        for row in range(2, tq + 2):
            if date_range_min < self.sheet[row][self.DATE].value < date_range_max:
                print(f"{Fore.BLUE + self.sheet[row][self.STATUS].value + Fore.RESET} "
                      f"{self.sheet[row][self.DATE].value} "
                      f"{self.sheet[row][self.ADDRESS].value} "
                      f"{self.sheet[row][self.TASK].value}"
                      f"{self.sheet[row][self.COMMENT].value}")

    def show_redcom_tasks(self):
        ...

    # todo Удаление дублей с подсчётом их кол-ва на ЖК и возвращением их кол-ва и процента от общего кол-ва заявок.
    #  Создание нового файла excel без дублей, с которым будут работать все функции
    # Прочитать книгу
    # отсортировать данные по ЖК и кол-ву дублей
    # записать данные в новую книгу без дублей
    # Сослаться на новую книгу во всех последующих функциях
