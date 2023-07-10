from types import NoneType

import shutil
import openpyxl
from colorama import Fore
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from Service import regulars


class OverportAnalizer:
    PATH = 'input_data/ads_journal.xlsx'

    DATE = 3
    ADDRESS = 4
    TASK = 5
    STATUS = 9
    COMMENT = 6

    SHEET_NAME = 'Without_Duplicates'
    ORIGINAL_SHEET_NAME = 'Лист 1'

    START_ROW = 2

    def __init__(self):

        self.wb: Workbook = openpyxl.load_workbook(self.PATH, read_only=False)

        if self.SHEET_NAME in self.wb.sheetnames:
            self.wb.remove(self.wb[self.SHEET_NAME])

        self.wb.create_sheet(self.SHEET_NAME, 0)
        self.wb.save(self.PATH)

        self.sheet: Worksheet = self.wb.active
        self.sheet_origin: Worksheet = self.wb[self.ORIGINAL_SHEET_NAME]

    def remove_duplicates(self):
        delete_row_count = 0
        for task in self.sheet_origin.iter_rows(min_row=self.START_ROW,
                                                max_row=self.sheet_origin.max_row,
                                                values_only=True):
            if task[self.COMMENT] is None:
                self.sheet.append(task)
            else:
                dublicate_flag = False
                for word in regulars.DUPLICATES:
                    if str(task[self.COMMENT].lower()).find(word) >= 0:
                        dublicate_flag = True
                        break
                if dublicate_flag:
                    delete_row_count += 1
                    continue
                self.sheet.append(task)
        self.wb.save(self.PATH)
        print(delete_row_count)

    def show_tasks(self, task_quantity=all, date_range_min='2020-01-01', date_range_max='2100-01-01'):
        # Определяем нужное кол-во тасков
        if task_quantity == all:
            tq = self.sheet.max_row
        else:
            tq = task_quantity

        # Определяем временно диапазон
        for row in range(2, tq + 2):
            if date_range_min < self.sheet[row][self.DATE].value < date_range_max:
                print(f"{Fore.BLUE + self.sheet[row][self.STATUS].value + Fore.RESET} "
                      f"{self.sheet[row][self.DATE].value} "
                      f"{self.sheet[row][self.ADDRESS].value} "
                      f"{self.sheet[row][self.TASK].value}"
                      f"{self.sheet[row][self.COMMENT].value}")

    # def show_redcom_tasks(self):
    #     ...
    #
    #
    # def show_reaction_time(self):
    #     ...
    #     # mix, max, middle
    #
    def show_all_task_by_objects(self):
        residential_task_count = {}
        for task in self.sheet_origin.iter_rows(min_row=self.START_ROW,
                                                max_row=5000,
                                                values_only=True):
            find_adress_flag = False
            for jk in regulars.RESIDENTIAL_COMPLEX:
                if find_adress_flag == True:
                    break
                else:
                    for adress in regulars.RESIDENTIAL_COMPLEX[jk]:
                        if str(task[self.ADDRESS]).find(adress) >= 0:
                            find_adress_flag = True
                            if jk in residential_task_count:
                                residential_task_count[jk] += 1
                                break
                            else:
                                residential_task_count[jk] = 1
                                break

        print(residential_task_count)
        summ = 0
        for v in residential_task_count.values():
            summ = summ + v
        print(summ)
        print(self.sheet_origin.max_row - 1)
        print(len(residential_task_count))
